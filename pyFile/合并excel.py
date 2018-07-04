#coding:utf8
'''
用于合并多个excel文件到一个文件当中。
'''
import sys
import os
import pandas as pd
import numpy
import numpy as np
from xlrd import xldate_as_tuple
from openpyxl import load_workbook
from xlrd import open_workbook
import  xlwt
from wx import *
def getFileList(path): #获取指定目录下的所有文件
    files=os.listdir(path)
    for file in files:
        if file[0]=='.': #跳过隐藏文件
            continue
        elif file.endswith(('xls','xlsx')): #定义需要读取的文件类型
            filelist.append(path+'/'+file) #MAC和win此处不同
    return filelist
def read_xls(address):
    rawdata1=open_workbook(address)
    rawdata=rawdata1.sheet_by_index(0)
    num=rawdata.nrows #行数
    #print("num ",num)
    for row in range(3,num):  #限制从第几行开始读取数据
        #print(rawdata.row_values(row))
        rdata=rawdata.row_values(row)  
        #print(rdata)
        datavalue.append(rdata)
    return datavalue  #datavalue用来放整个文件的单元格内容
def get_function(path):
    global filelist #定义全局存储文件名的列表，定义在调用函数内部最前面。
    filelist=[] #存放文件名
    if '.' in path:  #判断addressPath中是文件夹还是单个文件
        filelist.append(path)
    else:
        getFileList(path) #是文件夹，调用获取文件名函数
    global len_of_all
    global datavalue
    datavalue=[] #存放每个文件的所有单元格数据
    i=0
    for item in filelist:
        if ('xlsx' in item) or ('xls' in item):
            try:
                read_xls(item)
                '''
                rawdata1=open_workbook(address)
                rawdata=rawdata1.sheet_by_index(0)
                matrix[i]=[0]*(rawdata.nrows-1)
                for m in range(rawdata.nrows-1):
                    matrix[i][m]=["0"]*rawdata.ncols
                num=rawdata.nrows #行数
                for row in range(1,num):
                    for k in range(0,rawdata.ncols):
                        matrix[i][j-1][k]=rawdata.cell(j,k).value
                '''
            except:
                continue
        i+=1
import traceback
class MyFrame(Frame):
    def __init__(self):
        Frame.__init__(self,None,-1,title="自动合并excel by XT",pos=(100,100),size=(500,300))
        panel=Panel(self,-1)
        global cb
        self.button1=Button(panel,-1,"打开文件夹",pos=(170,60),size=(100,40))
        self.button2=Button(panel,-1,"RUN",pos=(170,120),size=(100,40))
        #cb2=RadioButton(panel,-1,label="Folder File",pos=(150,60))
        self.button3=Button(panel,-1,"写入文件",pos=(170,180),size=(100,40))
        self.button1.Bind(EVT_BUTTON,self.getMyPath)  
        self.button2.Bind(EVT_BUTTON,self.run_file)
        self.button3.Bind(EVT_BUTTON,self.get_write_path)
        #StaticText(panel,-1,"输出",pos=(70,200))
        #text_output=TextCtrl(panel,-1,pos=(100,200),style=TE_MULTILINE |  TE_READONLY,size=(250,200),) #多行显示&只读
        #self.__TextBox2=text_output
        StaticText(panel,-1,"Write By XT",pos=(390,220))
        self.InitUI() 
    def onClick(self,event):
        global cb
        cb=event.GetEventObject().GetLabel()
        #print "on:",event.GetEventObject().GetLabel()
        if str(event.GetEventObject().GetLabel())=="Folder File":
            cb=2
        else:
            cb=1
    def get_write_path(self,event):
        dlg = DirDialog(self,u"选择文件夹",style=DD_DEFAULT_STYLE)  
        if dlg.ShowModal() == ID_OK:  
            global Topath #定义全局变量Path
            Topath=dlg.GetPath()
            Topath=str(Topath)
            self.write_to_excel(Topath)
    def write_to_excel(self,toPath):
        Excel_file = xlwt.Workbook() 
        sheet = Excel_file.add_sheet('sheet0')
        styleBlueBkg = xlwt.easyxf('pattern: pattern solid, fore_colour green') #单元格颜色，淡蓝色
        outtxt=['15281283988', '18259229913', '15117896648', '15086161689', u'13023948469', '13688428317', '13377276210', '13573915109', '17355109152', u'13275861333', '15100210381', '13997686698', u'15077806789', '15285859343', '18814876184', '13626472211']

        #print(datavalue)
        try:
            print(len(datavalue))
            '''
            for i in range(len(datavalue)):  #这段代码是专用。
                tmp=datavalue[i][0]
                if str(tmp) in outtxt:
                    sheet.write(i,0,tmp,styleBlueBkg)
                else:
                    sheet.write(i,0,datavalue[i][0])
                for j in range(1,len(datavalue[i])):
                    c=datavalue[i][j]
                    sheet.write(i,j,c)
            '''
            
            
            for i in range(len(datavalue)): #这段代码是通用
                #print(len(datavalue[i]))
                for j in range(len(datavalue[i])):
                    c=datavalue[i][j]
                    sheet.write(i,j,c)
            
            
            Excel_file.save(str(toPath)+'/output.xlsx')
            dial=MessageDialog(None,"合并"+str(fileNum)+"个excel完成！")
            dial.ShowModal()
        except:
            traceback.print_exc()
            dial=MessageDialog(None,"可能文件太大，合并出错！")
            dial.ShowModal()
    def run_file(self,event):  #输出识别号码
        global matrix
        matrix=[None]*fileNum
        get_function(Path)    #调用识别文件函数
        dial=MessageDialog(None,"处理完成！")
        dial.ShowModal()
    def InitUI(self):    #自定义的函数,完成菜单的设置  
        menubar = MenuBar()        #生成菜单栏  
        filemenu = Menu()        #生成一个菜单  
        qmi1 = MenuItem(filemenu,1, "help")     #生成一个help菜单项  
        qmi2 = MenuItem(filemenu,2, "Quit")  #quit项，id设为2，在bind中调用
        filemenu.Append(qmi1)            #把菜单项加入到菜单中  
        filemenu.Append(qmi2)  
        menubar.Append(filemenu, "&File")        #把菜单加入到菜单栏中  
        self.SetMenuBar(menubar)            #把菜单栏加入到Frame框架中  
        self.Bind(EVT_MENU, self.OnQuit, id=2)    #给菜单项加入事件处理，id=2  
        self.Show(True)        #显示框架  
    def OnQuit(self, e):    #自定义函数　响应菜单项　　  
        self.Close()
    def getMyPath(self,event):
        #print "final cb:",cb
        #self.__TextBox2.Clear()  #清除文本框
        global Path
        dlg=DirDialog(self,"选择文件夹",style=DD_DEFAULT_STYLE)
        if dlg.ShowModal() == ID_OK:  
            Path=dlg.GetPath()
            #self.__TextBox1.SetLabel(Path) #设置textbox内容为文件内容
            global fileNum
            fileNum=0
            for lists in os.listdir(Path):                     # 统计文件数量
                sub_path = os.path.join(Path, lists)
                #print(sub_path)
                if os.path.isfile(sub_path):
                    fileNum = fileNum+1  
            return Path #文件夹路径
        dlg.Destroy() 
if __name__ == "__main__":
    app = App()    #创建应用的对象
    myframe = MyFrame()    #创建一个自定义出来的窗口
    #myframe.Center()#正中间显示
    myframe.Show()    #这两句一定要在MainLoop开始之前就执行    
    app.MainLoop()

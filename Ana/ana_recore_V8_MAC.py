#--coding:utf-8--
import sys
import os
import pandas as pd
import json
import numpy
import numpy as np
import time
from xlrd import xldate_as_tuple
import datetime
from datetime import datetime,date
from openpyxl import load_workbook
from xlrd import open_workbook
from wx import *

other_cell_phone1 = []
other_cell_phone=[]
other_cell_phone1 = []
other_cell_phone2=[]
start_time1=[]
start_time=[]
def getFileList(path): #获取指定目录下的所有文件
    #filelist=[] #存放文件名
    files=os.listdir(path)
    for file in files:
        if file[0]=='.': #跳过隐藏文件
            continue
        elif file.endswith(('json','csv','txt','xls','xlsx')): #定义需要读取的文件类型
            filelist.append(path+'/'+file)
    #print filelist
    return filelist
#path='/Users/xutao/Downloads/Python/data_set/phonedata_zhang/'
#path=Path
def select(address):
    #print "hello1"
    #print "Get:",num_Get
    try:
        if num_Get:
            num_get=int(num_Get) #若有定义输入，则用定义，否则默认为5个
    except:
        num_get=5
    start_hour=[]  #存放日期中的小时
    start_date=[]  #存放日-day
    start_date1=[] #存放月日,比如10月2号，以10.02数字存放
    start_weekday=[] #存放星期几
    year=datetime.now().year  #获取当前年份
    month=datetime.now().month #获取当前月份
    for item in start_time:
        #print type(item)
    #for i in range(len(calls)):
        if '/' in str(item): #判断日期格式
            if str(item).count('/')==2 and str(item).count(':')==2:
                start_weekday.append(datetime.strptime(item, "%Y/%m/%d %H:%M:%S").weekday()+1) #将日期转换为星期
                item=time.strptime(item, "%Y/%m/%d %H:%M:%S")  #将日期字符串转换为日期格式
                start_hour.append(item.tm_hour)  #提取出时间中的小时
                start_date.append(item.tm_mday) #提取日期
                start_date1.append(float(item.tm_mon)+float(1.0*(item.tm_mday)/100))
            elif str(item).count('/')==2 and str(item).count(':')==1:
            #print "2"
                start_weekday.append(datetime.strptime(item, "%Y/%m/%d %H:%M").weekday()+1) #将日期转换为星期
                item=time.strptime(item, "%Y/%m/%d %H:%M")  #将日期字符串转换为日期格式
                start_hour.append(item.tm_hour)  #提取出时间中的小时
                start_date.append(item.tm_mday) #提取日期
                start_date1.append(float(item.tm_mon)+float(1.0*(item.tm_mday)/100))

            elif str(item).count('/')==1 and str(item).count(':')==2:
                item='2017/'+str(item) #如果日期当中缺少年份，自动添加年份
                start_weekday.append(datetime.strptime(item, "%Y/%m/%d %H:%M:%S").weekday()+1) #将日期转换为星期
                item=time.strptime(item, "%Y/%m/%d %H:%M:%S")  #将日期字符串转换为日期格式
                start_hour.append(item.tm_hour)  #提取出时间中的小时
                start_date.append(item.tm_mday) #提取日期
                start_date1.append(float(item.tm_mon)+float(1.0*(item.tm_mday)/100))
            elif str(item).count('/')==1 and str(item).count(':')==1:
                item='2017/'+str(item)
                start_weekday.append(datetime.strptime(item, "%Y/%m/%d %H:%M").weekday()+1) #将日期转换为星期
                item=time.strptime(item, "%Y/%m/%d %H:%M")  #将日期字符串转换为日期格式
                start_hour.append(item.tm_hour)  #提取出时间中的小时
                start_date.append(item.tm_mday) #提取日期
                start_date1.append(float(item.tm_mon)+float(1.0*(item.tm_mday)/100))
            else:
                dial=MessageDialog(None,"时间数据格式/有误，请参照File-help！")
                dial.ShowModal()
                exit()
        else:
            if str(item).count('-')==2 and str(item).count(':')==2:
                start_weekday.append(datetime.strptime(item, "%Y-%m-%d %H:%M:%S").weekday()+1) #将日期转换为星期
                item=time.strptime(item, "%Y-%m-%d %H:%M:%S")  #将日期字符串转换为日期格式
                start_hour.append(item.tm_hour)  #提取出时间中的小时
                start_date.append(item.tm_mday) #提取日期
                start_date1.append(float(item.tm_mon)+float(1.0*(item.tm_mday)/100))
            elif str(item).count('-')==2 and str(item).count(':')==1:
                start_weekday.append(datetime.strptime(item, "%Y-%m-%d %H:%M").weekday()+1) #将日期转换为星期
                item=time.strptime(item, "%Y-%m-%d %H:%M")  #将日期字符串转换为日期格式
                start_hour.append(item.tm_hour)  #提取出时间中的小时
                start_date.append(item.tm_mday) #提取日期
                start_date1.append(float(item.tm_mon)+float(1.0*(item.tm_mday)/100))
            elif str(item).count('-')==1 and str(item).count(':')==2:
                item='2017-'+str(item)
                start_weekday.append(datetime.strptime(item, "%Y-%m-%d %H:%M:%S").weekday()+1) #将日期转换为星期
                item=time.strptime(item, "%Y-%m-%d %H:%M:%S")  #将日期字符串转换为日期格式
                start_hour.append(item.tm_hour)  #提取出时间中的小时
                start_date.append(item.tm_mday) #提取日期
                start_date1.append(float(item.tm_mon)+float(1.0*(item.tm_mday)/100))
            elif str(item).count('-')==1 and str(item).count(':')==1:
                item='2017-'+str(item)
                start_weekday.append(datetime.strptime(item, "%Y-%m-%d %H:%M").weekday()+1) #将日期转换为星期
                item=time.strptime(item, "%Y-%m-%d %H:%M")  #将日期字符串转换为日期格式
                start_hour.append(item.tm_hour)  #提取出时间中的小时
                start_date.append(item.tm_mday) #提取日期
                start_date1.append(float(item.tm_mon)+float(1.0*(item.tm_mday)/100))
            else:
                dial=MessageDialog(None,"时间数据格式-有误，请参照File-help！")
                dial.ShowModal()
                exit()
    #start_year_day.append(item.tm_year,item.tm_mon,item.tm_mday)
    #print start_min[i]start_hm.append(float(start_hour[i])+float(1.0*start_min[i]/60)) 
    #print len(start_weekday)

    #用day_count统计记录中出现的日期，每个日期只记录一次
    day_count=[]
    for item in start_date1:
        if(item in day_count):
            pass
        else:
            day_count.append(item)
   # print("记录中出现天数：",len(day_count) ) #打印出日期天数
    #统计每个时间段内的通话次数:
    #print min_hour,max_hour #测试用
    #print cal_usetime#测试用
    #检查other_phone_call中的不同电话号码数：
    #print other_cell_phone
    count=0
    dic={} #字典，存放号码和对应次数
    def dic_relation():
        for item in other_cell_phone:
            if(item in dic.keys()):
                dic[item]+=1
            else:
                dic[item]=1
    #max_call_time= max(dic.values())
    dic_relation()
    #print dic
    #print len(dic)
    #print start_date
    #建立词典，一个号码若每天都打电话，则标记为1，否则标记为比例：
    dic_phone_day={}
    dic_phone_work={}#统计工作日天数
    dic_phone_fev={}#节假日
    #先讲全部置一。
    for item in other_cell_phone:
        if(item in dic_phone_day.keys()):
            pass
        else:
            dic_phone_day[item]=1
    for item in other_cell_phone:
            dic_phone_work[item]=1
    for item in other_cell_phone:
            dic_phone_fev[item]=1
    #print dic_phone_work
    #统计每个号码的工作日电话数
    i=0
    #print(len(start_weekday),len(other_cell_phone),len(start_time))
    for item in dic_phone_work.keys():
        i=0
        j=0
        f=0
        for call in other_cell_phone:
            if (item==call):
                if(start_weekday[i]>=1 and start_weekday[i]<=5):
                    j+=1
                elif(start_weekday[i]>=6 and start_weekday[i]<=7):
                    f+=1
                else:
                    pass
            else:
                pass
            i+=1
        dic_phone_work[item]=j
        dic_phone_fev[item]=f
    #print dic_phone_fev   #测试
    dic_phone_ratio={} #存放通话通话比例
    def input_mm():#自定义输入上下限
        global Min
        Min=input("请输入双休-工作日通话次数比例下限(-1~+1):") 
        if Min<-1 or Min >1:
     #       print("最小值输入错误，请重新输入:")
            return input_mm()
        global Max
        Max=input("请输入双休-工作日通话次数比例上限(-1~+1):")
        if Max<-1 or Max >1:
      #      print("最大值输入错误，请重新输入:")
            return input_mm()  
        if Min>=Max:
       #     print("最小值大于最大值，请重新输入:" ) 
            return input_mm()
    #input_mm()
    #print dic_phone_work
    #print dic_phone_fev
    for item in dic_phone_day.keys():
        if item in dic_phone_ratio.keys():
            pass
        else:
            if len(item)>=8:
                dic_phone_ratio[item]=float(1.0*(dic_phone_fev[item]-dic_phone_work[item])/(dic_phone_fev[item]+dic_phone_work[item]))
    #print Min,Max #测试
    #建立根据比例筛选出的号码词典：
    dic_phone_select1={}
    for day_Gap in range(10,35,5):
        for item in dic_phone_ratio.keys():
            if dic[item]>=(len(day_count)/day_Gap):
                dic_phone_select1[item]=dic_phone_ratio[item]
        if len(dic_phone_select1)>num_get:
            break
    #sorted(dic_phone_select.iteritems(), key = lambda asd:asd[1], reverse=True)

    #print dic_phone_select

    #print "阈值为:",i,"num_get:",num_get
    #print "初步筛选个数:",len(dic_phone_select1)
    #print "按照节假日-工作日通话比例排序后："
    dic_phone_select={}
    i=0
    for item in sorted(dic_phone_select1.iteritems(), key = lambda asd:asd[1], reverse=True) :
        if i<num_get: #取前10个
            dic_phone_select[item[0]]=dic_phone_ratio[item[0]]  #按照从大到小排序输出,python2.7中sorted之后出来的item是元祖格式，提取号码需要item[0],python3不是这样。
            i+=1
        else:
            break
    #如下代码作用：统计出所有号码通话天数，即若某天有通话(无论一天中通话多少次)，则加一，
    #print("1")
    for item in dic_phone_day.keys():
        j=0
        tmp=[]
        for call in other_cell_phone:
        #print call
            if (item!=call):
                pass
                j+=1
            else:                                                                                                                                  
                if(start_date1[j] in tmp):
                    pass
                else:
                    tmp.append(start_date1[j])
                j+=1
        dic_phone_day[item]=len(tmp)
    #print("最多通话天数：",max(dic_phone_day.values()))
    most_related_num=max(dic_phone_day.items(),key=lambda x:x[1])[0]
    #print("最多通话天数号码:",most_related_num)
    #print dic_phone_day
    #print dic_phone_day_sort      
    #print dic_phone_day
    #print len(start_date)  #1173 
    #print max_call_time #159
    weekday={  #转换星期
    1:'星期一',
    2:'星期二',
    3:'星期三',
    4:'星期四',
    5:'星期五',
    6:'星期六',
    7:'星期日',
    }
    #rawdata1=pd.read_csv('/Users/xutao/Desktop/副本 通话详单采集1.csv') #MAC
    #rawdata11=[] #写入数组中
    #for i in range(len(rawdata1)):
    #   rawdata11.append(rawdata1['other_cell_phone'][i])
    #print rawdata11 #测试
    #global final_select
    #final_select=[]
    #print "筛选出的号码通话记录:"
    tmparray=[]
    for i in range(len(other_cell_phone)):
        if(start_hour[i]>=16   and start_weekday[i]>=5 and start_weekday[i]<=7): #and len(other_cell_phone[i])>8 )# and other_cell_phone[i][0]!='0'): #过滤掉0开始的号码):
            if(other_cell_phone[i] in dic_phone_select.keys()):  
            #print other_cell_phone[i],":通话开始时间:",start_hour[i],", 星期几通话:",weekday[start_weekday[i]],", 通话日期:",start_date1[i],"号"
                if(other_cell_phone[i] in tmparray):
                    pass
                else:
                    tmparray.append(other_cell_phone[i])
    #elif(other_cell_phone[i]=='18917009353'): #输出某个号码的通话记录
        #print other_cell_phone[i],":通话开始时间:",start_hour[i],"点, 通话时长:","s, 星期几通话:",weekday[start_weekday[i]],", 通话日期:",start_date1[i],"号,location:"
   # print(tmparray)
   # print("筛选出的号码为:")
   # print(most_related_num)
    if most_related_num in dic_phone_select:
        #print("ha :",most_related_num)
        i=0
        for item in sorted(dic_phone_select, key = lambda asd:asd[1], reverse=True):
            #print(item)
            if i>=num_get:
                break
            elif item in tmparray:
         #       print(i+1,":",item)
                final_select.append(item)
                i+=1
            else:
                pass
    else:
        i=0
        #print("1 :",most_related_num)
        final_select.append(most_related_num)
        for item in sorted(dic_phone_select, key = lambda asd:asd[1], reverse=True):
            #print("item",item)
            if i>num_get:
                break
            elif item in tmparray:
          #      print(i+2,":",item)
                final_select.append(item)
                i+=1
            else:
                pass
    filelength.append(len(final_select))  #保存当final_select的长度，也即识别出的号码的个数
    judgelist.append(1)  #若在处理过程中能走到这一步，这说明没有出错，置1.
    #print final_select
    #print filelength
def read_csv(address):
    rawdata=pd.read_csv(address,skip_blank_lines=True) #参数为去除空行
    #a=rawdata.iloc[:,1]
    #print rawdata
    #print len(rawdata)
    if 'start_time' or 'begin_time' or '起始时间' in rawdata.columns:
    #=='start_time':
        if '起始时间' in rawdata.columns:
            start_time1=rawdata['起始时间']
        elif 'begin_time' in rawdata.columns:
            start_time1=rawdata['begin_time']
        elif 'start_time' in rawdata.columns:
            start_time1=rawdata['start_time']
        else:
            judgelist.append(0) #处理出错，置0
            string=str(address)+"数据格式有误，请参照File-help！"
            dial=MessageDialog(None,string)
            #dial=MessageDialog(None,"csv数据格式有误，请参照File-help！")
            dial.ShowModal()
            exit()
            #print("No found:start_time !please make sure again.")
    if 'other_cell_phone' or 'opsite_phone' or '对方号码' in rawdata.columns:
        if 'other_cell_phone' in rawdata.columns:
            other_cell_phone1=rawdata['other_cell_phone']
        elif 'opposite_phone'in rawdata.columns:
            other_cell_phone1=rawdata['opposite_phone']
        elif'对方号码' in rawdata.columns:
            other_cell_phone1=rawdata['对方号码']
        else:
            #print("No found:other_cell_phone !please make sure again.")
            judgelist.append(0) #处理出错，置0
            string=str(address)+"数据格式有误，请参照File-help！"
            dial=MessageDialog(None,string)
            #dial=MessageDialog(None,"csv数据格式有误，请参照File-help！")
            dial.ShowModal()
            exit()
        #os._exit()
    #重新将数据转换为str:
    for i in range(len(start_time1)):
        #print start_time1[i]
        start_time.append(str(start_time1[i])) 
        other_cell_phone2.append(str(other_cell_phone1[i]))
    #整理电话号码的格式：   
    for i in range(len(other_cell_phone2)): #
        if other_cell_phone2[i][0]=='1' :
            #use_time.append(time2se(call['use_time']))#通话时长
            other_cell_phone.append(other_cell_phone2[i])
        elif len(other_cell_phone2[i])>=13: #加86开头的手机号
            other_cell_phone.append(other_cell_phone2[i][-11:]) #对方号码
        else:
            other_cell_phone.append(other_cell_phone2[i][-8:]) #对方号码
    select(address)
def read_json(address):
    with open(address,'r') as f: #ubuntu
        mobile = json.load(f)
    calls = mobile
    use_time = []
    i=0
    try:
        if  'other_cell_phone' in calls[0]:
            ocp='other_cell_phone'
        elif 'opposite_phone' in calls[0]:
            ocp='opposite_phone'
        elif '对方号码' in calls[0]:
            ocp='对方号码'
        if 'start_time' in calls[0]:
            st='start_time'
        elif 'begin_time' in calls[0]:
            st='begin_time'
        elif '开始时间' in calls[0]:
            st='开始时间'
    except:
        judgelist.append(0) #处理出错，置0
        string=str(address)+"数据格式有误，请参照File-help！"
        dial=MessageDialog(None,string) #报错同时报错文件名字和路径
        #dial=MessageDialog(None,"json数据格式有误，请参照File-help！")
        dial.ShowModal()
        exit()
    for call in calls:   
        #print call['other_cell_phone'][0] 
        #if(call['other_cell_phone']and call['start_time']):
            #print type(str(call['start_time'])) #测试数据类型用
            #print call[ocp]
            if (ocp in call) and (st in call): #这行代码是为了防止json中的通话记录中掺杂有短信等其他信息
                if len(call[st])<1:
                    continue
                elif len(str(call[ocp]))<1: #跳过缺失数据
                    continue
                elif len(str(call[ocp]))==0:
                    other_cell_phone1.append(str(call[oct]))
                elif str(call[ocp])[0]=='1' :
                    other_cell_phone1.append(str(call[ocp])) #对方号码
                elif len(str(call[ocp]))>=13:
                    other_cell_phone1.append(str(call[ocp])[-11:]) #对方号码
                else:
                    other_cell_phone1.append(str(call[ocp])[-8:]) #对方号码
                start_time.append(str(call[st])) #将通话开始时间加入
            else:
                pass
    #整理电话号码格式：
    #print type(other_cell_phone1[0][0])
    for i in range(len(other_cell_phone1)): #
        #print len(other_cell_phone1[i])
        if len(other_cell_phone1[i])==0:
            #print i,"backspace"
            other_cell_phone.append(other_cell_phone1[i])
        elif other_cell_phone1[i][0]=='1' :
            #print "haha"
            other_cell_phone.append(other_cell_phone1[i])
        elif len(other_cell_phone1[i])>=13: #加86开头的手机号
            other_cell_phone.append(other_cell_phone1[i][-11:]) #对方号码
        else:
            other_cell_phone.append(other_cell_phone1[i][-8:]) #对方号码  
    select(address)
def read_txt(address):
    rawdata=open(address,'r')
    i=0
    a=[] #c存放第一行的列名
    for line in rawdata:
        if i==1: #默认第二行开始存储通话数据
            a=line.split(',')
            #if '/' or '-' in a: # a[0].decode('utf-8'):
            #print "1",len(a)
            for j in range(len(a)): #查找指定列名所在的列下标
                if (('-' in str(a[j]))or('/' in str(a[j]))): #判断日期所在列数
                    date_index=j #保存日期的列下标
                    #print j,str(a[j]) #测试
                elif  str(a[j]).isdigit(): #默认全为数字组成的字符串为电话号码
                    phone_index=j
                    #print phone_index #测试
                else:
                    pass
            break
                
        else:
            i+=1
    #print phone_index,date_index
    try:
        if (int(phone_index)+  int(date_index))>0:
            pass
    except:
        judgelist.append(0) #处理出错，置0
        string=str(address)+"数据格式有误，请参照File-help！"
        dial=MessageDialog(None,string)
        dial.ShowModal()
        exit()
    #date_index为日期列下表，phone_index为电话列下标
    i=0
    for line in rawdata:#开始转存数据：
        if len(line)<10: #跳过空行
            continue
        data_line=line.split(',') #txt默认以','分隔数据
        if i==0:
            pass #第一行为列名，跳过
            i+=1
        else:#从第二行开始保存数据
            start_time.append(data_line[date_index])
            other_cell_phone2.append(data_line[phone_index])
    #整理电话号码格式：
    for i in range(len(other_cell_phone2)): #
        if other_cell_phone2[i][0]=='1' :
        #use_time.append(time2se(call['use_time']))#通话时长
            other_cell_phone.append(other_cell_phone2[i])
        elif len(other_cell_phone2[i])>=13: #加86开头的手机号
            other_cell_phone.append(other_cell_phone2[i][-11:]) #对方号码
        else:
            other_cell_phone.append(other_cell_phone2[i][-8:]) #对方号码
    select(address)
    #print other_cell_phone #测试用
def read_xls(address):
    rawdata1=open_workbook(address)
    rawdata=rawdata1.sheet_by_index(0)
    #print rawdata.cell(1,0).ctype
    
    #取消了在excel中识别日期列的功能：if ctype==3那几行，默认第一列为日期，防止意外发生。
    try:
        for i in range(0,rawdata.ncols): #默认第一行为列名
            ctype=rawdata.cell(1,i).ctype
            #print ctype    #ctype : 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
            value1=rawdata.cell(1,i).value
            #if ctype==3:  #ctype为3表示为excel中的日期
                #date = xldate_as_tuple(value1,0)  
                #cell =datetime(*date) #转换为时间格式
                #if ('/' in str(cell)) or( '-' in str(cell)):
                #if ('/' in str(rawdata.cell_value(5,i).encode('utf-8'))) or( '-' in str(rawdata.cell_value(5,i).encode('utf-8'))):
                    #date_index=i #判断日期和号码的列下标
                    #print "index:",i
            #else:
                #pass
            if ctype==2 and len(str(value1))>4:
                #if (str(rawdata.cell_value(5,i).encode('utf-8'))).isdigit(): #默认全为数字组成的字符串为电话号码
                    phone_index=i
                    #print "phone:",i
                    #print str(int(rawdata.cell(1,i).value))
            if ctype==1:
                #if ('/' in str(rawdata.cell_value(5,i).encode('utf-8'))) or( '-' in str(rawdata.cell_value(5,i).encode('utf-8'))):
                    #date_index=i #判断日期和号码的列下标
                if (str(rawdata.cell_value(5,i).encode('utf-8'))).isdigit(): #默认全为数字组成的字符串为电话号码
                    phone_index=i
    except:
        judgelist.append(0) #处理出错，置0
        string=str(address)+"数据格式有误，请参照File-help！"
        dial=MessageDialog(None,string)
        #dial=MessageDialog(None,"Excel数据格式有误，请参照File-help！")
        dial.ShowModal()
        exit()
    #print date_index,phone_index #测试
    #print rawdata.nrows
    date_index=0 #默认第一列为日期，
    ctype=rawdata.cell(1,date_index).ctype #先获取日期的ctype
    for i in range(1,rawdata.nrows):
        #ctype=rawdata.cell(1,i).ctype
        #print ctype
        if rawdata.cell(i,date_index).value=="": #跳过空行
            #print i
            continue
        else:
            if ctype==3:  #若为3，则用datetime模块处理日期
                date1=rawdata.cell(i,date_index).value
                date2 = xldate_as_tuple(date1,0) 
                date3=datetime(*date2)
                phone1=str(int(rawdata.cell(i,phone_index).value))
                #print i,date3,phone1
                #   else:#从第二行开始保存数据
                start_time.append(str(date3))
                other_cell_phone2.append(str(phone1))
            else:
                start_time.append(str(rawdata.cell_value(i,date_index)))
                other_cell_phone2.append(str(rawdata.cell_value(i,phone_index)))


    for i in range(len(other_cell_phone2)): #
        if other_cell_phone2[i][0]=='1' :
        #use_time.append(time2se(call['use_time']))#通话时长
            other_cell_phone.append(other_cell_phone2[i])
        elif len(other_cell_phone2[i])>=13: #加86开头的手机号
            other_cell_phone.append(other_cell_phone2[i][-11:]) #对方号码
        else:
            other_cell_phone.append(other_cell_phone2[i][-8:]) #对方号码
        #print type(rawdata.cell_value(i,0))
    #print rawdata
    #print "hello"
    select(address)
#根据address的文件名，调用对应解析函数
def get_function(path):
    global filelist #定义全局存储文件名的列表，定义在调用函数内部最前面。
    global filelength #全局变量，存放每个文件识别出的号码长度
    global judgelist  #存放每个文件夹内的文件在处理的时候是否出错，若出错，置0，否则置1.
    #global filelength1
    filelength=[]
    #filelength1=[]
    filelist=[] #存放文件名
    judgelist=[]
    if '.' in path:  #判断addressPath中是文件夹还是单个文件
        filelist.append(path)
    else:
        getFileList(path) #是文件夹，调用获取文件名函数
    global final_select #全局数组final_select放在for外面，用来可以保存文件夹的多个文件。
    final_select=[]
    #print "hello0"
    #print filelist[0]
    for item in filelist:
        #print "hello 0-0"
        global other_cell_phone#定义全局数组，这样每次调用一个文件，都会产生新数组，防止不同文件相互叠加
        global other_cell_phone1
        global other_cell_phone2
        global start_time
        global start_time1
        other_cell_phone1 = []  
        other_cell_phone=[]
        other_cell_phone2=[]
        start_time1=[] 
        start_time=[]
        #print "hello 0-1"
        #print("文件:",item)
        if 'csv' in item:
        #t=threading.Thread(target=read_csv(item))
        #t.start()
            try:
                read_csv(item)
            except:
             #   print("error with:",item)
                continue
        #time.sleep(5)
        elif 'json' in item:
            try:
                read_json(item)
            except:
                continue
        #time.sleep(5)
        elif 'txt' in item:
            try:
                read_txt(item)
            except:
                continue
        #time.sleep(5)
        elif ('xlsx' in item) or ('xls' in item):
            try:
                read_xls(item)
            except:
                continue
        #time.sleep(5)
        else:
            dial=MessageDialog(None,"不支持的文件格式！")
            dial.ShowModal()

class MyFrame(Frame):
    def __init__(self):
        Frame.__init__(self,None,-1,title="通话记录分析",pos=(100,100),size=(800,600))
        panel=Panel(self,-1)
        global cb
        #cb=1
        self.button1=Button(panel,-1,"打开文件",pos=(370,100))
        self.button2=Button(panel,-1,"RUN",pos=(370,200),size=(100,40))
        self.button3=Button(panel,-1,"写入文件",pos=(370,320),size=(100,40))
        self.button4=Button(panel,-1,"确定",pos=(300,150),size=(60,20))

        cb1=RadioButton(panel,-1,label="Sigle file",pos=(70,60)) #定义文件读取方式，单个or文件夹形式
        self.Bind(EVT_RADIOBUTTON,self.onClick,cb1)

        cb2=RadioButton(panel,-1,label="Folder File",pos=(150,60))
        self.Bind(EVT_RADIOBUTTON,self.onClick,cb2)

        self.button1.Bind(EVT_BUTTON,self.getMyPath)
        
        StaticText(panel,-1,"文件",pos=(70,100))
        text=TextCtrl(panel,-1,pos=(100,100),size=(250,20))
        self.__TextBox1=text

        StaticText(panel,-1,"请输入想要输出的号码个数：\n(默认为5-6个)",pos=(70,150))
        text_input=TextCtrl(panel,-1,pos=(230,150),size=(60,20))
        self.__TextBox3=text_input
        print text_input

        self.button4.Bind(EVT_BUTTON,self.getInput)

        self.button2.Bind(EVT_BUTTON,self.run_file)

        StaticText(panel,-1,"输出",pos=(70,200))
        text_output=TextCtrl(panel,-1,pos=(100,200),style=TE_MULTILINE |  TE_READONLY,size=(250,200),) #多行显示&只读
       
        self.__TextBox2=text_output

        self.button3.Bind(EVT_BUTTON,self.get_write_path)
        StaticText(panel,-1,"Version:1.0",pos=(700,510))
        self.InitUI() 
    
    def onClick(self,event):
        global cb
        cb=event.GetEventObject().GetLabel()
        #print "on:",event.GetEventObject().GetLabel()
        if str(event.GetEventObject().GetLabel())=="Folder File":
            cb=2
        else:
            cb=1
        #print "on after:",cb,type(cb)
    def getInput(self,event): #得到输入的号码个数
        self.__TextBox2.Clear()
        #self.__TextBox3.GetValue()
        global num_Get
        num_Get=self.__TextBox3.GetValue()
        #print num_Get
    def run_file(self,event):  #输出识别号码

       # print("it is worked")
        #print("a")
        get_function(Path)    #调用识别文件函数
        #print("b")
        global fileLength  #全局变量，因为在写入时还要用到。
        fileLength=[] #存储每个文件号码的个数
        fileLength.append(filelength[0]) #先存第一个，以为下面的for循环相当于跳过了第一个，所以在for之前先append第一个数据
        for i in range(len(filelength)-1):
            fileLength.append(filelength[i+1]-filelength[i])
        #print fileLength
        if cb==1:  #单个文件的输出方式
            i=1
            for item in final_select:
                #print item
                self.__TextBox2.AppendText(str(i)+":")
                self.__TextBox2.AppendText(str(item)+'\n')
                i+=1
                #self.__TextBox2.AppendText("\n")
        else: #文件夹文件的输出方式，先输出文件名，再输出文件号码个数的号码
            i=0 #控制final_select
            j=0 #控制judgelist
            f=0 #控制filelist
            #print judgelist
            #print len(final_select)
            #print len(filelist)
            for item in filelist:#还没写完。。。。。。！！！！！！！！！！！！
                self.__TextBox2.AppendText(str(item)+'\n')#输出文件名
                if judgelist[j]==1: #当前文件没有错误，输出。
                    #print fileLength[f]
                    for k in range(int(fileLength[f])): #输出当前文件的号码长度个数，若当前文件在处理时出错，
                        self.__TextBox2.AppendText(str(final_select[i])+'\n')
                        i+=1 #finalselct后移
                    j+=1 #judgelist加1，
                    f+=1 #filelist加1
                elif judgelist[j]==0:  #当前文件有误，跳过。
                    pass
                    j+=1 #文件错误，judgelist加1，但filelist不动
    def InitUI(self):    #自定义的函数,完成菜单的设置  
        menubar = MenuBar()        #生成菜单栏  
        filemenu = Menu()        #生成一个菜单  
        qmi1 = MenuItem(filemenu,1, "help")     #生成一个help菜单项  
        qmi2 = MenuItem(filemenu,2, "Quit")  #quit项，id设为2，在bind中调用
        filemenu.AppendItem(qmi1)            #把菜单项加入到菜单中  
        filemenu.AppendItem(qmi2)  

        menubar.Append(filemenu, "&File")        #把菜单加入到菜单栏中  
        self.SetMenuBar(menubar)            #把菜单栏加入到Frame框架中  
        self.Bind(EVT_MENU, self.OnQuit, id=2)    #给菜单项加入事件处理，id=2  
        self.Bind(EVT_MENU, self.help_window, id=1)  #help窗口
        self.Show(True)        #显示框架  

    def OnQuit(self, e):    #自定义函数　响应菜单项　　  
        self.Close()

    def help_window(self,event): #定义help窗口
        dial=MessageDialog(None,"txt、csv、xls文件请尽量使用如下数据格式：\n第一行的列名：\n通话日期:start_time,\n通话location:place,"\
        +"\n通话类型:init_type,\n对方号码：other_cell_phone,\n通话持续时间:use_time\n数据格式："\
        +"\n通话日期:****/**/** **:**:**,\n对方号码:************,"\
        +"\n通话持续时长:**:**:**\n"\
        +"json文件格式:\n[\n  {\n   'start_time':****/**/** **:**:**\n   'other_cell_phone':***********\n   'use_time':**:**:**\n"\
        +"   ......\n   }\n   ......\n]\n不一样的格式可能导致无法读取！请尽量删去套餐优惠一列，以防止出现错误。",pos=(10,10)) #测试用
        dial.ShowModal()

    def get_write_path(self,event):
        dlg = DirDialog(self,u"选择文件夹",style=DD_DEFAULT_STYLE)  
        if dlg.ShowModal() == ID_OK:  
            #print "1"
            global Topath #定义全局变量Path
            Topath=dlg.GetPath()
            
            #print Topath
            print cb
            try:
                if cb==1:
                    filename=(Path.split("/")[-1]).split(".")[0] #获取读取的文件名，来作为写入文件名，MAC和window此处不同。
                    #print filename
                    with open(Topath+"/"+str(filename)+'.txt','a+') as f: #MAC和window此处不同。
                        f.write(str(filename)+":"+"\n")
                        for item in final_select:
                            f.write(str(item)+'\n') #将筛选号码写入txt
                        f.write('\n')
                    dial=MessageDialog(None,"写入成功！")
                    dial.ShowModal()
                    return Topath #文件夹路径  
                else:
                    print Topath
                    with open(Topath+'/selectNum.txt','a+') as fw: #竟然不能命名为f！！调试了半天没有发现哪里错误。。。MAC和window此处不同。
                        fw.write(str(Topath)+':'+'\n')
                        fw.flush()
                        i=0 #控制final_select
                        j=0 #控制judgelist
                        f=0 #控制filelist
                        print filename
                        print fileLength
                        print final_select
                        print judgelist
                        for item in filelist:
                            filename=(item.split("/")[-1]).split(".")[0] #MAC和window此处不同。
                            fw.write(str(filename)+':'+'\n') 
                            #self.__TextBox2.AppendText(str(item)+"\n")#输出文件名
                            if judgelist[j]==1: #当前文件没有错误，输出。
                                for k in range(int(fileLength[f])): #输出当前文件的号码长度个数，若当前文件在处理时出错，
                                    fw.write(str(final_select[i])+"\n")
                                #self.__TextBox2.AppendText(str(final_select[i])+"\n")
                                    i+=1 #finalselct后移
                                fw.write('\n')
                                j+=1 #judgelist加1，
                                f+=1 #filelist加1
                            elif judgelist[j]==0:  #当前文件有误，写入error。
                                fw.write('error!'+'\n'+'\n')
                                j+=1 #文件错误，judgelist加1，但filelist不动
                        dial=MessageDialog(None,"写入成功！")
                        dial.ShowModal()


            except:
                #print "error!"
                dial=MessageDialog(None,"error")
                dial.ShowModal()
        dlg.Destroy() 

    def getMyPath(self,event):
        #print "final cb:",cb
        self.__TextBox2.Clear()  #清除文本框

        if cb==2:
            dlg=DirDialog(self,"选择文件夹",style=DD_DEFAULT_STYLE)
            if dlg.ShowModal() == ID_OK:  
                global Path #定义全局变量Path
                Path=dlg.GetPath()
                self.__TextBox1.SetLabel(Path) #设置textbox内容为文件内容
                return Path #文件夹路径       
            dlg.Destroy() 
        else:
            dlg = FileDialog(self,u"选择文件",style=DD_DEFAULT_STYLE)  
            if dlg.ShowModal() == ID_OK:  
                global Path #定义全局变量Path
                Path=dlg.GetPath()
                self.__TextBox1.SetLabel(Path) #设置textbox内容为文件内容
                return Path #文件夹路径       
            dlg.Destroy() 

if __name__ == "__main__":
    global Path
    global final_select
    app = App()    #创建应用的对象
    myframe = MyFrame()    #创建一个自定义出来的窗口
    #myframe.Center()#正中间显示
    myframe.Show()    #这两句一定要在MainLoop开始之前就执行    
    app.MainLoop()